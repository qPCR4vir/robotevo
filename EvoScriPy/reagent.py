# Copyright (C) 2014-2019, Ariel Vina Rodriguez ( arielvina@yahoo.es )
#  distributed under the GNU General Public License, see <http://www.gnu.org/licenses/>.
#
# author Ariel Vina-Rodriguez (qPCR4vir)
# 2014-2019
"""
`Reagent` - a fundamental concept
=================================

A `Reagent` is a fundamental concept in RobotEvo programming. It makes possible to define a protocol in a natural way,
matching what a normal laboratory's protocol indicates.
Defines a named homogeneous liquid solution, the wells it occupy, the initial amount needed to run the protocol
(auto calculated), and how much is needed per sample, if applicable. It is also used to define samples,
intermediate reactions and products. It makes possible a robust tracking of all actions and a logical error
detection, while significantly simplifying the  programming of non trivial protocols.

todo: implement units for volume, concentration, etc.

Main classes and functions:
^^^^^^^^^^^^^^^^^^^^^^^^^^^

Abstract information classes:
-----------------------------

 - :py:class:`MixComponent`: like an item in some table summarizing components of some Mix.
 - :py:class:`PreMixComponent`: like an item in some table summarizing components of some PreMix.
 - :py:class:`Primer`: like an item in some table summarizing primer sequences, synthesis, etc.
 - :py:class:`PrimerMixComponent`: like an item in a table describing Primer Mixes for some PCRs.
 - :py:class:`PrimerMix`: like a table describing Primer Mixes for some PCRs
 - :py:class:`PCRMasterMix`: like an item in some table summarizing PCR Master Mixes for some PCR experiment
 - :py:class:`PCReaction`: like an item in some table summarizing reactions in a PCR experiment
 - :py:class:`PCRexperiment`: like an item in some table summarizing PCR experiments


Robot classes:
--------------

 - :py:class:`Reagent`: homogeneous liquid solution in some wells
 - :py:class:`MixReagent`: a Reagent composed of other Reagents
 - :py:class:`Dilution`: A mix defined by the dilutionn of one or more components with a diluent.
 - :py:class:`PreMixReagent`: A pre-MixReagent of otherwise independent reagents
 - :py:class:`PrimerReagent`: Manipulate a Primer Reagent on a robot.
 - :py:class:`PrimerMixReagent`: Manipulate a Primer-Mix Reagent on a robot.
 - :py:class:`PCRMasterMixReagent`: Manipulate a PCR Master-Mix Reagent on a robot.
 - :py:class:`PCReactionReagent`: Define a PCR reaction in a well or tube.
 - :py:class:`PCRexperimentRtic`: Organize a PCR setup on a robot.

 todo: set in code when a Reagent is an stock solution: it will be "prepared" following an small sequence of
 instructions, mostly independent of the current protocol. Examples: buffer TE 1x, primers, primers mixes, etc.
 (one critical difference is that the total volume to prepare is set independentl< of the number samples, volume or
 concentration, etc. need in the current protocol).
 This in contrast to solution specially prepared for the current protocol, like the "extraction" pre-mix in an RNA
 extraction protocol or a PCR "master mix"

"""

__author__ = 'qPCR4vir'


import logging
from pathlib import Path

import EvoScriPy.labware as lab

def_reagent_excess = 4
def_mix_excess = 8


class Reagent:
    """
    A Reagent is a fundamental concept in RobotEvo programming.
    It makes possible to define a protocol in a natural way, matching what a normal
    laboratory's protocol indicates. Defines a named homogeneous liquid solution,
    the wells it occupy, the initial amount needed to run the protocol (auto calculated),
    and how much is needed per sample, if applicable. It is also used to define samples,
    intermediate reactions and products. It makes possible a robust tracking
    of all actions and a logical error detection, while significantly simplifying
    the programming of non trivial protocols.
    """

    current_protocol = None                                               # to (auto)register a list of reagents
    use_minimal_number_of_aliquots = True

    def __init__(self,
                 name          : str,
                 labware       : (lab.Labware, str, [])    = None,
                 wells         : (int, [int], [lab.Well])  = None,
                 num_of_aliquots: int                      = None,
                 minimize_aliquots: bool                   = None,
                 def_liq_class : (str, (str, str))         = None,
                 volpersample  : float                     = 0.0,   #: todo move to PreMixComponent
                 num_of_samples: int                       = None,  #: todo move to Exp? to prepare?
                 single_use    : float                     = None,  #: todo move to MixComponent
                 excess        : float                     = None,  #: todo move to MixComponent
                 initial_vol   : (float, list)             = 0.0,
                 min_vol       : float                     = 0.0,
                 fill_limit_aliq: float                    = 100,
                 concentration : float                     = None   # todo implement use. Absolut vs. relative? Units?
                 ):
        """
        This is a named set of aliquots of an homogeneous solution.
        Put a reagent into labware wells, possible distributed into aliquots and set the amount to be used for each sample,
        if applicable.
        Each reagent is added to a list of reagents of the worktable were the labware is.
        The specified excess in % will be calculated/expected. A default excess of 4% will be assumed
        if None is indicated.
        A minimal needed volume will be calculated based on either the number of samples
        and the volume per sample to use or the volume per single use. This can be forced setting min_vol.
        A minimal number of replicas (wells, aliquots) will be calculated based on the minimal volume,
        taking into account the maximum allowed volume per well and the excess specified. Aliquots will be filled not more
        than the percent of the well volumen indicated by fill_limit_aliq.

        :param name:            Reagent name. Ex: "Buffer 1", "forward primer", "IC MS2"
        :param labware:         :py:class:`labware.Labware` or his label in the worktable; if None will be deduced from `wells`.
        :param volpersample:    how much is needed per sample, if applicable, in uL
        :param single_use:      Not a "per sample" multiple use? Set then here the volume for one single use
        :param wells:           or offset to begging to put replica. If None will try to assign consecutive wells in labware
        :param num_of_aliquots: def min_num_of_replica(), number of replicas
        :param def_liq_class:   the name of the Liquid class, as it appears in your own EVOware database.
                                instructions.def_liquidClass if None
        :param excess;          in %
        :param initial_vol:     is set for each replica. If default (=None) is calculated als minimum.
        :param min_vol:         force a minimum volume to be expected or prepared.
        :param fill_limit_aliq:    maximo allowed volume in % of the wells capacity
        :param num_of_samples:  if None, the number of samples of the current protocol will be assumed
        :param minimize_aliquots:  use minimal number of aliquots? Defaults to `Reagent.use_minimal_number_of_aliquots`,
                                   This default value can be temporally change by setting that global.

        """
        self.concentration = concentration
        logging.debug("Creating Reagent " + name)

        self.user_min_vol = min_vol or 0.0
        self.need_vol = 0.0  #: calculated volume needed during the execution of the protocol
        if labware is None:
            if isinstance(wells, lab.Well):
                labware = wells.labware
            elif isinstance(wells, list) and isinstance(wells[0], lab.Well):
                labware = wells[0].labware
        self.labwares = labware if isinstance(labware, list) else [labware]
        for idx, lw in enumerate(self.labwares):
            if isinstance(lw, str):
                self.labwares[idx] = Reagent.current_protocol.worktable.get_labware(label=lw)
            else:
                assert isinstance(lw, lab.Labware), "No labware defined for Reagent " + str(self)
        self.labware = self.labwares[0]

        # add self to the list of reagents of the worktable were the labware is.
        worktable = None
        assert isinstance(self.labware.location.worktable,   lab.WorkTable)                                   # todo temporal
        if (    isinstance(self.labware,                     lab.Labware)
            and isinstance(self.labware.location,            lab.WorkTable.Location)
            and isinstance(self.labware.location.worktable,  lab.WorkTable)):

            worktable = self.labware.location.worktable
        else:
            if Reagent.current_protocol:
                worktable = Reagent.current_protocol.worktable                         # todo temporal

        assert name not in worktable.reagents, "The reagent " + name + " was already in the worktable"
        worktable.reagents[name] = self
        ex = def_reagent_excess if excess is None else excess

        self.fill_limit_aliq = 1.0 if fill_limit_aliq is None else fill_limit_aliq / 100.0
        self.excess = 1.0 + ex/100.0
        self.def_liq_class = def_liq_class
        self.name = name
        self.volpersample = volpersample
        self.components = []                                # todo reserved for future use ??
        if minimize_aliquots is not None:
            self.minimize_aliquots = minimize_aliquots
        else:
            self.minimize_aliquots = Reagent.use_minimal_number_of_aliquots
        if single_use:
            assert not volpersample, str(name) + \
                                     ": this is a single use-reagent. Please, don't set any volume per sample."
            # if num_of_samples is None:
            #     num_of_samples = 1
            assert num_of_samples is None, "this is a single use-reagent, don't set num_of_samples " + str(num_of_samples)

            self.need_vol = single_use
        else:
            if volpersample and num_of_samples is None:
                num_of_samples = Reagent.current_protocol.num_of_samples or 0

        self.min_num_aliq = self.min_num_of_replica(num_of_samples=num_of_samples)

        # coordinate: min_num_aliq, wells, num_of_aliquots, initial_vol. todo fine-tune warnings.
        if num_of_aliquots is None:
            num_of_aliquots = self.min_num_aliq
        assert num_of_aliquots >= self.min_num_aliq, ("too few wells (" + str(len(wells)) + ") given for the minimum "
                                                      + "number of replica needed (" + str(self.min_num_aliq) + " )")

        if isinstance(wells, list):                         # use len(pos) to correct num_of_aliquots
            assert len(wells) >= self.min_num_aliq, ("too few wells (" + str(len(wells)) + ") given for the minimum "
                                                     + "number of replica needed (" + str(self.min_num_aliq) + " )")
            if num_of_aliquots < len(wells):
                logging.warning("WARNING !! putting more replica of " + name + " to fit the initial volume list provided")
                num_of_aliquots = len(wells)
            assert not num_of_aliquots > len(wells), ("too few wells (" + str(len(wells)) + ") given for the desired "
                                                      + "number of replicas (" + str(num_of_aliquots) + " )")
            if isinstance(initial_vol, list):
                assert num_of_aliquots == len(initial_vol)
            if num_of_aliquots > self.min_num_aliq:                              # todo revise !! for PreMixReagent components
                logging.warning("WARNING !! You may be putting more wells replicas (" + str(len(wells)) + ") of " + name
                                + " that the minimum you need(" + str(self.min_num_aliq) + " )")

        if not isinstance(initial_vol, list):
            initial_vol = [initial_vol] * num_of_aliquots

        if num_of_aliquots < len(initial_vol):
            logging.warning("WARNING !! putting more replica of " + name + " to fit the initial volume list provided")
            num_of_aliquots = len(initial_vol)
            if num_of_aliquots > self.min_num_aliq:                                      # todo revise !! for PreMixReagent components
                logging.warning("WARNING !! You may be putting more inital volumens values (" + str(num_of_aliquots) + ") of " + name
                                + " that the minimum number of replicas you need(" + str(self.min_num_aliq) + " )")

        try:
            self.Replicas   = self.labware.put(self, wells, self.min_num_aliq if self.minimize_aliquots else num_of_aliquots)
        except lab.NoFreeWells as er:
            logging.warning("No free wells: " + str(er))
            for lwre in self.labwares:
                if lwre is self.labware:
                    continue
                try:
                    self.Replicas = lwre.put(self, wells,
                                                     self.min_num_aliq if self.minimize_aliquots else num_of_aliquots)
                    self.labware = lwre
                    break
                except lab.NoFreeWells as er:
                    logging.warning("No free wells: " + str(er))
            else:
                er = lab.NoFreeWells(labware=self.labware, error=' to put ' + str(num_of_aliquots)
                                                                 + ' aliquots of reagent - ' + str(self))
                logging.warning(str(er))
                raise er

        self.pos        = self.Replicas[0].offset                                   # ??

        self.initial_vol = None
        self.init_vol(num_samples=num_of_samples, initial_vol=initial_vol)            # put the minimal initial volume
        self.include_in_check = True
        logging.debug("Created Reagent " + str(self))

    def min_num_of_replica(self, num_of_samples: int = None) -> int:
        """
        A minimal number of replicas (wells, aliquots) will be calculated based on the minimal volume,
        taking into account the maximum allowed volume per well and the excess specified.
        :param num_of_samples:
        :return:
        """
        min_vol = self.min_vol(num_of_samples)
        if min_vol is None:
            logging.warning("Calculating min_num_of_replica for " + str(self) + " before setting any volume.")
            min_vol = 0.0
        return int(min_vol / (self.labware.type.max_vol * self.fill_limit_aliq)) + 1

    @staticmethod
    def set_reagent_list(protocol):
        Reagent.current_protocol = protocol                    # ??

    @staticmethod
    def stop_reagent_list():
        Reagent.current_protocol = None

    def __str__(self):
        return "{name:s}".format(name=self.name)

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.Replicas or '-') + ']'

    def min_vol(self, num_samples=None, volume: float = None, add_volume: float = None) -> float:
        """
        The minimal needed volume will be calculated based on either the number of samples
        and the volume per sample to use, or  as an accumulation of all the volume "requested" by add_volume from the
        preparation of other components of the protocol(todo or the volume per single use.)???

        :param num_samples: total number of samples to prepare. Use this only for "per sample" reagents, in which case
                            this can be set to None to use the protocols' number of samples.
        :param volume: just set this volume as the minimal required
        :param add_volume: add this to the minimum
        :return: the minimal needed volume calculated (this is also "saved" in member self.need_vol)
        """

        if self.volpersample:
            assert not (volume or add_volume), "This is a 'per sample' reagent, " + str(self) + \
                                               "don't try to set or change the needed volume directly"
            num_samples = num_samples or Reagent.current_protocol.num_of_samples or 0
            self.need_vol = self.volpersample * num_samples * self.excess
        else:
            assert not num_samples, "This is not a 'per sample' reagent, " + str(self) + \
                                    "don't try to set the number of samples to calculate the needed volume"
            if volume:
                assert not add_volume
                self.need_vol = volume * self.excess
            elif add_volume:
                assert not volume
                if self.need_vol is None:
                    self.need_vol = 0.0
                self.need_vol += add_volume * self.excess

        return self.need_vol

    def init_vol(self, num_samples=None, initial_vol=None):
        """
        To initialize the among of reagent in each well. First put what the user inform he had put, then
        put additionally the minimum the protocol need.
        :param num_samples:
        :param initial_vol: what the user inform he had put on each aliquot well.
        :return:
        """
        if initial_vol is not None:
            assert isinstance(initial_vol, list)
            self.initial_vol = []
            for w, v in zip(self.Replicas, initial_vol):  # todo and the rest wells? 0?
                w.vol = v if v else 0
                assert w.labware.type.max_vol >= w.vol, 'Excess initial volume for ' + str(w)
                self.initial_vol.append(w.vol)
        else:
            for w, v in zip(self.Replicas, self.initial_vol):  # todo and the rest wells? 0?
                w.vol = v if v else 0
                assert w.labware.type.max_vol >= w.vol, 'Excess initial volume for ' + str(w)

        self.put_min_vol(num_samples)

    def put_min_vol(self, num_samples=None):          # todo create num_of_aliquots if needed !!!!
        """
        Force you to put an initial volume of reagent that can be used to distribute into samples,
        aspiring equal number of complete doses for each sample from each replica,
        except the firsts replicas that can be used to aspirate one more dose for the last/rest of samples.
        That is: all replica have equal volume (number) of doses or the firsts have one more dose
        :param num_samples:
        :return:
        """
        if self.need_vol is None:
            logging.warning("Puting put_min_vol for " + str(self) + " before setting any volume.")
            self.need_vol = 0.0

        need_vol = max(self.need_vol, self.user_min_vol, 0.0)
        if self.volpersample:
            if num_samples is None:
                num_samples = Reagent.current_protocol.num_of_samples
            if not num_samples and not need_vol:
                return
            v_per_sample = self.volpersample * self.excess
            num_samples = max(num_samples, (need_vol + 0.5) // v_per_sample )
            replicas=len(self.Replicas)
            for i, w in enumerate(self.Replicas):
                v = v_per_sample * (num_samples + replicas - (i + 1)) // replicas
                if v > w.vol:  w.vol += (v-w.vol)
                assert w.labware.type.max_vol >= w.vol, 'Add one more replica for ' + w.reagent.name
        else:
            vr = need_vol * self.excess / len(self.Replicas)
            for w in self.Replicas:
                if vr > w.vol:  w.vol += (vr-w.vol)
                assert w.labware.type.max_vol >= w.vol, 'Add one more replica for ' + w.reagent.name

    def autoselect(self, maxTips=1, offset=None, replicas = None):
        # todo revise !!!!!!!!! we know the wells = Replicas
        return self.labware.autoselect(offset or self.pos, maxTips, len(self.Replicas) if offset is None else 1)

    def select_all(self):
        return self.labware.selectOnly([w.offset for w in self.Replicas])


class NoReagentFound(Exception):
    def __init__(self, reagent_name: str, error: str):
        self.reagent_name = reagent_name
        Exception.__init__(self, "No reagent named " + str(reagent_name) + error)  # todo redaction


class MixComponent:
    """
    Represent abstract information, like an item in some table summarizing components of some MixReagent.
    todo: introduce diluent? - final_conc == None ? final_conc == init_conc ?
    """

    def __init__(self,
                 name: str,
                 id_: str = None,
                 init_conc: float = None,
                 final_conc: float = None,
                 volume: float = None):
        """

        :param id_:
        :param name:
        :param init_conc:  todo Really?? explore None for lyo
        :param final_conc:
        """
        self.id = id_
        self.name = name
        self.init_conc = init_conc
        self.final_conc = final_conc
        self.volume = volume

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'


class MixComponentReagent:
    """
    Components of some MixReagent.
    This is not just a Reagent, but some "reserved" volume of some Reagent.

    """

    def __init__(self,
                 reagent: Reagent,
                 volume: float = None):
        """

        :param reagent:
        :param volume:
        """
        self.reagent = reagent
        self.vol = 0.0
        self.excess = 1.0
        self.volume(volume)

    def __str__(self):
        return self.reagent.name

    def __repr__(self):
        return (self.reagent.name or '-') + '[' + str(self.volume or '-') + ']'

    def volume(self, vol: float = None, excess: float = None):  # todo put_min_vol in reagent?
        old_vol = self.vol * self.excess
        if vol is None and excess is None:
            return old_vol

        if excess is not None:
            self.excess = excess

        if vol is not None:
            self.vol = vol

        new_vol = self.vol * self.excess

        self.reagent.min_vol(add_volume=new_vol - old_vol)
        self.reagent.init_vol()
        return new_vol


class MixReagent(Reagent):
    """
    A Reagent composed of other Reagents, that the robot may prepare.
    This is just a mix of fixed volume of the mixed components, with in turn fixes the total volume
    of this reagent.
    """

    def __init__(self,
                 name          : str,  # todo introduce abstract info class MixReagent and rename this MixReagent?
                 labware       : (lab.Labware, str, [])    = None,
                 wells         : (int, [int], [lab.Well])  = None,
                 components    : [MixComponentReagent]     = None,
                 num_of_aliquots: int                      = None,
                 minimize_aliquots: bool                   = None,
                 def_liq_class : (str, (str, str))         = None,
                 excess        : float                     = None,
                 initial_vol   : float                     = None,
                 min_vol       : float                     = 0.0,  # todo ??
                 fill_limit_aliq: float                    = 100,
                 concentration : float                     = None,
                 volpersample  : float                     = None
                 ):
        """

        :param name:
        :param labware:
        :param wells:
        :param components:
        :param num_of_aliquots:
        :param minimize_aliquots:
        :param def_liq_class:
        :param excess:
        :param initial_vol:
        :param min_vol:
        :param fill_limit_aliq:
        :param concentration:
        """
        ex = def_mix_excess if excess is None else excess

        Reagent.__init__(self, name,
                         labware,
                         wells= wells,
                         num_of_aliquots= num_of_aliquots,
                         def_liq_class = def_liq_class,
                         excess = ex,
                         initial_vol = initial_vol,
                         fill_limit_aliq= fill_limit_aliq,
                         concentration=concentration,
                         min_vol=min_vol,
                         minimize_aliquots=minimize_aliquots,
                         volpersample=volpersample)

        if initial_vol == 0.0:  # the user signal it will be prepared
            vol = 0.0
            for comp in components:
                assert isinstance(comp, MixComponentReagent)
                vol += comp.volume(ex)
                # comp.reagent.put_min_vol()
            self.min_vol(add_volume=-vol)
        self.components = components  #: list of reagent components
        # self.init_vol()

    def __str__(self):  # todo ?
        return "{name:s}".format(name=self.name)

    def __repr__(self):  # todo ?
        return (self.name or '-') + '[' + str(self.Replicas or '-') + ']'

    def make(self, protocol, volume=None):      # todo deprecate?
        if self.Replicas[0].vol is None:   # ????
            self.put_min_vol(volume)
            assert False
        protocol.make_mix(self, volume)


class DilutionComponentReagent(MixComponentReagent):
    """
    Components of some Dilution.

    """

    def __init__(self,
                 reagent: Reagent,
                 dilution: float = None,
                 final_conc: float = None):
        """

        :param reagent:
        :param dilution:
        :param final_conc:
        """
        self.dilution = dilution
        self.final_conc = final_conc
        if dilution:
            assert not final_conc
        else:
            assert reagent.concentration
            self.dilution = reagent.concentration / final_conc
        assert self.dilution >= 1.0
        # vol = final_vol / dilution
        MixComponentReagent.__init__(self, reagent)  # , volume=vol

    def __str__(self):
        return self.reagent.name

    def __repr__(self):
        return (self.reagent.name or '-') + '[1/' + str(self.dilution or '-') + ']'

    def volume(self, vol: float = None, excess: float = None):
        if vol is not None:
            if not self.dilution:
                assert self.reagent.concentration
                self.dilution = self.reagent.concentration / self.final_conc
            assert self.dilution >= 1.0
            vol = vol / self.dilution
            MixComponentReagent.volume(self, vol=vol)
        return self.vol


class Dilution(MixReagent):
    """
    A Reagent mix composed of others - to be diluted - Reagents and a diluter, that the robot may prepare.
    """

    def __init__(self,
                 name: str,  # todo introduce abstract info class MixReagent and rename this MixReagent?
                 diluent: Reagent,
                 labware: (lab.Labware, str, []) = None,
                 wells: (int, [int], [lab.Well]) = None,
                 components: [DilutionComponentReagent] = None,
                 num_of_aliquots: int = None,
                 minimize_aliquots: bool = None,
                 def_liq_class: (str, (str, str)) = None,
                 excess: float = None,
                 initial_vol: float = 0.0,
                 min_vol: float = 0.0,
                 fill_limit_aliq: float = 100,
                 concentration: float = None
                 ):
        """

        :param diluent:
        :param name:
        :param labware:
        :param wells:
        :param components: [list of :py:class:`DilutionComponentReagent` 's]
        :param num_of_aliquots:
        :param minimize_aliquots:
        :param def_liq_class:
        :param excess:
        :param initial_vol:
        :param min_vol:
        :param fill_limit_aliq:
        :param concentration:
        """
        ex = def_mix_excess if excess is None else excess
        vol = 0.0
        self.diluent = None
        assert isinstance(components, list)
        for comp in components:
            if comp is diluent:
                self.diluent = diluent
                continue
            assert isinstance(comp, DilutionComponentReagent)
            vol += comp.volume(min_vol - initial_vol, ex)
        if self.diluent is None:
            self.diluent = diluent
            components.append(diluent)
        diluent.min_vol(add_volume= (min_vol - initial_vol)*ex - vol)

        if initial_vol is None:
            initial_vol = 0.0

        MixReagent.__init__(self, name,
                            labware,
                            components=components,
                            wells=wells,
                            num_of_aliquots=num_of_aliquots,
                            def_liq_class=def_liq_class,
                            excess=ex,
                            initial_vol=initial_vol,
                            fill_limit_aliq=fill_limit_aliq,
                            concentration=concentration,
                            min_vol=min_vol,
                            minimize_aliquots=minimize_aliquots)

        # self.components = components  #: list of reagent components
        # self.init_vol()

    def reserve(self, need_vol: float):
        vol = 0.0
        for comp in self.components:
            assert isinstance(comp, DilutionComponentReagent)
            vol += comp.volume(need_vol)

    def __str__(self):  # todo ?
        return "{name:s}".format(name=self.name)

    def __repr__(self):  # todo ?
        return (self.name or '-') + '[' + str(self.Replicas or '-') + ']'

    def make(self, protocol, volume=None):  # todo deprecate?
        if self.Replicas[0].vol is None:  # ????
            self.put_min_vol(volume)
            assert False
        protocol.make_mix(self, volume)


class PreMixComponent(MixComponent):
    """
    Represent abstract information, like an item in some table summarizing components of some PreMixReagent.
    An special case of MixComponent, for which volume is calculated on the basis of "number of samples"
    and volume_per_sample
    """

    def __init__(self,
                 name: str,
                 volpersample: float,
                 id_: str = None,
                 init_conc: float = None,
                 final_conc: float = None,
                 volume: float = None):
        """

        :param id_:
        :param name:
        :param init_conc:  todo Really??
        :param final_conc:
        """
        self.volpersample = volpersample
        MixComponent.__init__(self,
                              id_=id_,
                              name=name,
                              init_conc=init_conc,
                              final_conc=final_conc,
                              volume=volume)

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'


class PreMixReagent(Reagent):  # todo rewrite to be a MixReagent
    """
    A pre-MixReagent of otherwise independent reagents to be pippeted together for convenience,
    but that could be pippeted separately.
    todo: make this a special case of MixReagent, for which everything ? is calculated on the basis of "number of samples"
    """

    def __init__(self,
                 name,
                 labware: (lab.Labware, list),
                 components,
                 pos = None,
                 num_of_aliquots = None,
                 initial_vol = None,
                 def_liq_class = None,
                 excess = None,
                 fill_limit_aliq =None,
                 num_of_samples = None  # todo here?
                 ):
        """

        :param name:
        :param labware:
        :param components: list of reagent components
        :param pos:
        :param num_of_aliquots:
        :param initial_vol:
        :param def_liq_class:
        :param excess:
        :param fill_limit_aliq:
        :param num_of_samples:
        """
        ex = def_mix_excess if excess is None else excess
        vol = 0.0
        for reagent in components:
            vol += reagent.volpersample
            reagent.excess += ex/100.0      # todo revise! best to calculate at the moment of making?
            reagent.put_min_vol(num_of_samples)

        if initial_vol is None:
            initial_vol = 0.0

        Reagent.__init__(self, name,
                         labware,
                         volpersample= vol,
                         wells= pos,
                         num_of_aliquots= num_of_aliquots,
                         def_liq_class = def_liq_class,
                         excess = ex,
                         initial_vol = initial_vol,
                         fill_limit_aliq= fill_limit_aliq,
                         num_of_samples = num_of_samples)

        self.components = components  #: list of reagent components
        # self.init_vol()

    def init_vol(self, num_samples=None, initial_vol=None):
        """
        update my self.volpersample from the already updated self.components.volpersample,
        possibly with updated num_samples and initial_vol
        WARNING !! call this only after the update of the components (if need)
        :param num_samples:
        :param initial_vol:
        """
        if self.components:  # todo ??
            self.volpersample = 0
            for reagent in self.components:
                self.volpersample += reagent.volpersample
            Reagent.init_vol(self, num_samples=0, initial_vol=initial_vol)
        pass
        # put volume in num_of_aliquots only at the moment of making  !!
        # Reagent.init_vol(self, num_samples)
        # self.put_min_vol(num_samples)

    def make(self, protocol, NumSamples=None):      # todo deprecate? not used
        if self.Replicas[0].vol is None:   # todo ????
            self.put_min_vol(NumSamples)   # todo ????
            assert False
        protocol.make_pre_mix(self, NumSamples)  # used directly?

    def compVol(self,index, NumSamples=None):
        NumSamples = NumSamples or Reagent.current_protocol.num_of_samples
        return self.components[index].min_vol(NumSamples)


class Reaction(Reagent):
    """
    todo: make this a MixReagent, with diluent too ?
    """
    def __init__(self,
                 name,
                 labware,
                 components: [Reagent] = None,
                 track_sample=None, # just one more component?
                 pos=None,
                 num_of_aliquots=1,
                 def_liq_class=None,
                 excess=None,
                 initial_vol=0):

        Reagent.__init__(self,
                         name,
                         labware,
                         single_use=0,
                         wells=pos,
                         num_of_aliquots=num_of_aliquots,
                         def_liq_class=def_liq_class,
                         excess=excess,
                         initial_vol=initial_vol)
        self.track_sample = track_sample


class Primer:
    """
    Represent abstract information, like an item in some table summarizing primer sequences, synthesis, etc.
    """

    ids = {}  #: connect each existing Primer ID with the corresponding Primer 's
    seqs = {}  #: connect each existing Primer sequence with the corresponding list of Primer synthesis
    names = {}  #: connect each existing Primer name with the corresponding list of Primer synthesis
    key_words = {}  #: connect each existing Primer key_word with the corresponding list of Primer 's
    ids_synt = {}  #: connect each existing Primer synthesis ID with the corresponding Primer

    next_internal_id = 0

    def __init__(self,
                 name: str,
                 seq: str,
                 proposed_stock_conc: float=100,  #: uM
                 id_: str=None,
                 prepared: float=None,
                 mass: float=None,  #: ug
                 nmoles: float=None,  #: nmoles
                 molec_w: float=None,  #: g/mol
                 mod_5p: str=None,
                 mod_3p: str=None,
                 id_synt: str=None,
                 kws: list=None,
                 diluent: str='TE 1x'):
        """

        :param name: the external "human readable name"
        :param seq: the nucleotide sequence
        :param proposed_stock_conc: typically 100 (100 uM, or ~100x more than in PCR)
        :param id_: the external ID, typically a number string aimed to be unique
        :param prepared: volume (typically in uL) already in solution or need to be prepared (=None) from
               the originally provided lyophilized primer?
        :param mass: how much was synthesised, ugr.
        :param nmoles: how much was synthesised, mmoles.
        :param molec_w: molecular weight, gr/mol
        :param mod_5p: chemical modification, like flourecent FAM, or biotin
        :param mod_3p: chemical modification, like BHQ1, etc.
        :param id_synt: the external synthesis number provided by the supplier firm (unique)
        :param kws: any collection of key-words the user want to attach to this primer
                   (name of targets, virus, gen, project, owner, etc.)
        :param diluent: name of the reagent used to dilute the lyophilized primer.
        """
        self.diluent = diluent
        self.prepared = prepared
        self.proposed_stock_conc = proposed_stock_conc
        self.mass = mass
        self.nmoles = nmoles
        self.molec_w = molec_w
        self.mod_5p = mod_5p
        self.mod_3p = mod_3p
        self.id_synt = id_synt
        self.name = name
        self.seq = seq
        self.id = id_
        self._internal_id = Primer.next_internal_id #: internal unique ID
        Primer.next_internal_id += 1
        Primer.names.setdefault(name, []).append(self)

        if id_ and id_ in Primer.ids:
            wrn = 'Duplicate primer ID: ' + str(id_)
            logging.warning(wrn)
            # raise Warning(wrn)
        Primer.ids.setdefault(id_, []).append(self)

        if not id_synt:
            id_synt = 'FakeID-' + str(self._internal_id)
        if id_synt in Primer.ids_synt:
            err = 'Duplicate primer synthesis ID: ' + str(id_synt)
            logging.error(err)
            raise Exception(err)
        Primer.ids_synt[id_synt] = self

        Primer.seqs.setdefault(seq, []).append(self)
        if isinstance(kws, list):
            for kw in kws:
                Primer.key_words.setdefault(kw, []).append(self)

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'

    @staticmethod
    def load_excel_list(file_name: Path = None):
        col = {'conc': 0,
               'id': 2,
               'prepared': 4,
               'name': 5,
               'moles': 7,
               'mass': 8,
               'seq': 15,
               'mol_w': 11,
               'mod_5p': 19,
               'mod_3p': 20,
               'ido': 22,
               'virus': 25,
               'diluent': 1
               }
        logging.debug("opening excel")
        import openpyxl

        if not file_name:
            file_name = Path('C:\Prog\exp\PCR fli.xlsx')

        if not file_name:
            from tkinter import filedialog
            file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                   defaultextension='fas',
                                                   title='Select HEV isolate subtyping deta')
            if not file_name:
                return

        logging.debug(file_name)

        wb = openpyxl.load_workbook(str(file_name))
        logging.debug(wb.sheetnames)

        ws = wb['Primers']
        first = True
        p = None
        for r in ws.iter_rows() :
            if first:
                diluent = r[col['diluent']].value
                first = False
            else:
                p=Primer(name=r[col['name']].value,
                         seq=r[col['seq']].value,
                         prepared=r[col['prepared']].value,
                         proposed_stock_conc=r[col['conc']].value,
                         id_=r[col['id']].value,
                         mass=r[col['mass']].value,
                         nmoles=r[col['moles']].value,
                         molec_w=r[col['mol_w']].value,
                         mod_5p=r[col['mod_5p']].value,
                         mod_3p=r[col['mod_3p']].value,
                         id_synt=r[col['ido']].value,
                         kws=[r[col['virus']].value],
                         diluent=diluent
                         )
        return p


class PrimerReagent(Dilution):
    """
    Manipulate a Primer Reagent on a robot.
    """
    excess = def_mix_excess


    def __init__(self,
                 primer: Primer,
                 primer_rack: (lab.Labware, list),
                 pos=None,
                 initial_vol=None,
                 diluent: Reagent=None,
                 PCR_conc=0.8,
                 stk_conc=100,
                 def_liq_class=None,
                 fill_limit_aliq=None,
                 excess=None):
        """
        Construct a robot-usable PrimerReagent from an abstract Primer.
        You can reuse "old" aliquots by passing primer.prepared volume > 0.
        If no  primer.prepared volume is passed, it will be prepared (diluted) from the lyophilized primer.

        :param primer: :py:class:`Primer`
        :param primer_rack: labware rack where to put this primer. If a list - provide first the preferred racks.
        :param pos: an optional position where to put the primer in the rack. If =None one position will be proposed
        :param initial_vol: in uL
        :param diluent: optionally the dilution Reagent to be used, if None - will be "deduced" from primer
        :param PCR_conc: optional, used by default, but normally overrides by the primer mix reagent.
        :param stk_conc:
        :param def_liq_class:
        :param fill_limit_aliq:
        :param excess: optional, % to prepare in excess
        """

        self.stk_conc = stk_conc
        self.primer = primer
        if not diluent:
            pass  # todo identify or create diluent Reagent

        Dilution.__init__(self,
                          name=primer.name,
                          diluent=diluent,
                          labware=primer_rack,  # or lab.stock, ??
                          wells=pos,
                          components=[diluent],
                          num_of_aliquots=1,
                          initial_vol=initial_vol or primer.prepared,
                          excess=excess or PrimerReagent.excess,
                          fill_limit_aliq=fill_limit_aliq,
                          def_liq_class=def_liq_class,
                          concentration=stk_conc)

        # todo prepare unprepared Primers  !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

    def __str__(self):
        return "primer " + self.primer.name

    def __repr__(self):
        return "primer " + (self.primer.name or '-') + '[' + str(self.primer.id or '-') + ']'

    def dilute(self):
        Reagent.current_protocol.dilute_primer(self, 1000 * self.primer.nmoles / self.stk_conc)


class PrimerMixComponent(MixComponent):
    """
    Represent abstract information, like an item in a table describing Primer Mixes for some PCRs.
    It can be a primer, another primer mix or the diluent
    """

    def __init__(self,
                 id_ = None,
                 name = None,
                 init_conc = None,
                 final_conc = None,
                 super_mix: bool = False):
        """

        :param id_:
        :param name:
        :param init_conc:
        :param final_conc:
        :param super_mix:
        """
        MixComponent.__init__(self, id_, name, init_conc, final_conc)
        self.super_mix = super_mix


class PrimerMix:
    """
    Represent abstract information, like a table describing Primer Mixes for some PCRs
    """

    ids = {}  #: connect each existing Primer mix ID with the corresponding PrimerMix
    names = {}  #: connect each existing Primer mix name with the corresponding list of PrimerMix
    key_words = {}  #: connect each existing Primer mix key_word with the corresponding list of PrimerMix
    super_mix = True

    next_internal_id = 0

    def __init__(self,
                 name,
                 id_ = None,
                 conc = 10.0,
                 prepared=None,
                 components = None,
                 ref_vol = None,
                 diluent = None,
                 kws = None,
                 super_mix = False):
        """

        :param name:
        :param id_:
        :param conc:
        :param prepared:
        :param components:
        :param ref_vol:
        :param diluent:
        :param kws:
        :param super_mix:
        """
        self.prepared = prepared
        self.super_mix = super_mix
        self.diluent = diluent
        self.name = name
        self.id = id_
        self.conc = conc
        self.components = components
        self.ref_vol = ref_vol

        self._internal_id = PrimerMix.next_internal_id
        PrimerMix.next_internal_id += 1

        PrimerMix.names.setdefault(name, []).append(self)
        PrimerMix.ids.setdefault(id_, []).append(self)
        if isinstance(kws, list):
            for kw in kws:
                PrimerMix.key_words.setdefault(kw, []).append(self)

    def __str__(self):
        return "primer mix " + self.name

    def __repr__(self):
        return "primer mix " + (self.name or '-') + '[' + str(self.id or '-') + ']'

    @staticmethod
    def load_excel_list(file_name: Path = None):
        col = {'conc': 2,
               'id': 0,
               'name': 1,
               'vol': 2,
               'prepared': 4,
               'final': 6,
               'virus': 13,
               'super_mix': 1
               }
        logging.debug("opening excel")
        import openpyxl

        if not file_name:
            file_name = Path('C:\Prog\exp\PCR fli.xlsx')

        if not file_name:
            from tkinter import filedialog
            file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                   defaultextension='fas',
                                                   title='Select HEV isolate subtyping deta')
            if not file_name:
                return

        logging.debug(file_name)

        wb = openpyxl.load_workbook(str(file_name))
        logging.debug(wb.sheetnames)

        ws = wb['Primer mix']  # todo argument

        # todo another dic
        no_l, header, name_l, vol_l, sep_l, table_h_l, primer_l, diluter_l = 0, 1, 2, 3, 4, 5, 6, 7

        line = no_l
        components = []

        diluent = 'TE 0,1 x'  # todo ????

        pmix = None
        id_ = None
        name = None
        conc = None
        ref_vol = None
        kws = None
        super_mix = False
        prepared = None

        for r in ws.iter_rows():

            if line == no_l:
                line += 1

            elif line == header:
                id_ = r[col['id']].value
                kws = [r[col['virus']].value]
                superm = r[col['super_mix']].value
                super_mix = ('SuperMix' == superm)
                line += 1

            elif line == name_l:
                # assert id == r[col['id']].value
                name = r[col['name']].value
                conc= r[col['conc']].value
                line += 1

            elif line == vol_l:
                ref_vol = r[col['vol']].value
                prepared = r[col['prepared']].value
                line += 1

            elif line == sep_l:
                line += 1

            elif line == table_h_l:
                line += 1

            elif line == primer_l:
                comp = PrimerMixComponent(id_ = r[col['id']].value,
                                          name = r[col['name']].value,
                                          init_conc = r[col['conc']].value,
                                          final_conc = r[col['final']].value,
                                          super_mix = ('"x"' == r[col['conc'] + 1].value))

                if comp.name == diluent:  # todo diluent always last?
                    components.append(comp)
                    pmix = PrimerMix(name=name,
                                     id_=id_,
                                     conc=conc,
                                     ref_vol=ref_vol,
                                     prepared=prepared,
                                     kws=kws,
                                     components=components,
                                     diluent=comp,
                                     super_mix=super_mix
                                     )
                    line = no_l
                    components = []

                    id_ = None
                    name = None
                    conc = None
                    ref_vol = None
                    kws = None
                    super_mix = False
                    prepared = None

                elif comp.final_conc and (comp.id or comp.name):
                    components += [comp]

        return pmix


class PrimerMixReagent(PreMixReagent):
    """
    Manipulate a Primer-MixReagent Reagent on a robot.
    """

    excess = def_mix_excess

    def __init__(self,
                 primer_mix: PrimerMix,
                 primer_mix_rack: (lab.Labware, list),
                 pos=None,
                 num_of_aliquots=None,
                 initial_vol=None,
                 def_liq_class=None,
                 excess=None,
                 fill_limit_aliq=None,
                 primer_rack: (lab.Labware, list) = None):
        """
        Construct a robot-usable PrimerMixReagent from an abstract PrimerMix.
        You can reuse "old" aliquots by passing primer_mix.prepared volume > 0.
        If no  primer_mix.prepared volume is passed, or if it is not sufficient,
        a set of primer reagents will be created.


        :param primer_mix:
        :param primer_mix_rack:
        :param pos:
        :param num_of_aliquots:
        :param initial_vol:
        :param def_liq_class:
        :param excess:
        :param fill_limit_aliq:
        :param primer_rack:
        """
        self.primer_mix = primer_mix

        initial_vol = initial_vol or self.primer_mix.prepared

        components = []
        vol = 0
        lw = primer_mix_rack[0] if isinstance(primer_mix_rack, list) else primer_mix_rack
        reagents = lw.location.worktable.reagents
        diluent_vol = self.primer_mix.ref_vol             # todo ??

        for component in self.primer_mix.components:
            assert isinstance(component, PrimerMixComponent)
            if component is self.primer_mix.diluent:
                vol_per_mix = diluent_vol
            else:
                vol_per_mix = self.primer_mix.ref_vol * component.final_conc / float(component.init_conc)
                diluent_vol -= vol_per_mix
            if component.name in reagents:  # for example an existing primer
                component_r = reagents[component.name]
                assert isinstance(component_r, Reagent)
                # if not abs(component_r.volpersample - vol_per_reaction) < 0.05:
                #    assert False
                # component_r.put_min_vol()
            else:
                if component.name in PrimerMix.names:

                    for prmx in PrimerMix.names[component.name]:
                        assert isinstance(prmx, PrimerMix)
                        if component.id == prmx.id:
                            c_primer_mix = prmx
                            break
                    else:
                        raise NoReagentFound(component.name, " as primer mix with ID:" + str(component.id)
                                                             + ". Alternatives are: " + str(PrimerMix.names[component.name]))

                    component_r = PrimerMixReagent(c_primer_mix,
                                                   primer_mix_rack=primer_mix_rack,
                                                   primer_rack=primer_rack)
                elif component.name in Primer.names:

                    for pr in Primer.names[component.name]:
                        assert isinstance(pr, Primer)
                        if component.id == pr.id:
                            c_primer = pr
                            break
                    else:
                        raise NoReagentFound(component.name, " as primer with ID:" + str(component.id)
                                                             + ". Alternatives are: " + str(Primer.names[component.name]))

                    component_r = PrimerReagent(c_primer,
                                                primer_rack=primer_rack)

            components.append(component_r)

        PreMixReagent.__init__(self,
                        primer_mix.name,
                        primer_mix_rack,
                        pos,
                        components,
                        num_of_aliquots=num_of_aliquots,
                        initial_vol=initial_vol,
                        excess=excess or PrimerMixReagent.excess)


    def __str__(self):
        return "primer mix " + self.primer_mix.name

    def __repr__(self):
        return "primer mix " + (self.primer_mix.name or '-') + '[' + str(self.primer_mix.id or '-') + ']'


class ExpSheet:
    def __init__(self,
                 file_name: Path,
                 page,
                 cell_rows,
                 sample_line,
                 num_col=12,
                 num_row=8,
                 ):

        self.file_name = file_name
        self.page = page
        self.num_col = num_col
        self.num_row = num_row
        self.cell_rows = cell_rows
        self.sample_line = sample_line
        name = Path(file_name).stem.split('.')
        self.id = name[0]
        self.title = '.'.join(name[1:])


class PCRMasterMix:
    """
    Represent abstract information, like an item in some table summarizing PCR Master Mixes for some PCR experiment
    """

    ids = {}  #: connect each existing PCR master mix ID with the corresponding PCRMasterMix
    names = {}  #: connect each existing PCR master mix name with the corresponding PCRMasterMix
    next_internal_id = 0

    def __init__(self,
                 name,
                 id_=None,
                 reaction_vol=25,  #: in uL
                 sample_vol=5,  #: in  uL
                 components=None,
                 diluent=None,
                 title=None):
        """

        :param name:
        :param id_:
        :param reaction_vol:
        :param sample_vol:
        :param components:
        :param title:
        """
        self.diluent = diluent
        self.name = name
        self.id = id_
        self.reaction_vol = reaction_vol
        self.sample_vol = sample_vol
        self.title = title
        self.components = components
        self._internal_id = PCRMasterMix.next_internal_id
        PCRMasterMix.next_internal_id += 1
        assert name not in PCRMasterMix.names
        PCRMasterMix.names[name] = self
        assert id_ not in PCRMasterMix.ids
        PCRMasterMix.ids[id_] = self

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'

    @staticmethod
    def load_excel_list(exp_sheet: ExpSheet, ws=None):
        col = {'conc': 16,
               'id': 14,
               'name': 14,
               'vol': 16,
               'final': 18,
               'sample_v': 16,
               'title': 15,
               'comp_name': 15
               }

        if ws is None:
            logging.debug("opening excel file")
            import openpyxl

            if not exp_sheet.file_name:        # deprecate
                file_name = Path('C:\Prog\exp\PCR fli.xlsx')

            if not exp_sheet.file_name:       # deprecate
                from tkinter import filedialog       # deprecate
                file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                       defaultextension='fas',
                                                       title='Select HEV isolate subtyping deta')
                if not file_name:
                    return

            logging.debug(exp_sheet.file_name)

            wb = openpyxl.load_workbook(str(exp_sheet.file_name))
            logging.debug(wb.sheetnames)

            ws = wb[exp_sheet.page or 'Druken']       # deprecate

        no_l, header, name_l, vol_l, sampl_l, table_h_l, comp_l, diluent_l = 0, 0, 0, 1, 2, 5, 6, 7
        diluent = 'H2O'   # ?
        line = no_l
        pmix = None
        id = None
        name = None
        vol_per_reaction = 25
        components = []
        sample_vol = 5  # uL
        title = None
        to_skeep = 1 + exp_sheet.num_row * exp_sheet.cell_rows
        skeeped = 0

        for r in ws.iter_rows():

            if skeeped < to_skeep:
                skeeped += 1
                continue

            if line <= name_l:
                name = r[col['name']].value
                if name:
                    title = r[col['title']].value
                    line += 1

            elif line == vol_l:
                id = r[col['id']].value
                vol_per_reaction = r[col['vol']].value
                line += 1

            elif line == sampl_l:
                sample_vol = r[col['sample_v']].value
                line += 1

            else:
                component = MixComponent(id_=r[col['id']].value,
                                         name=r[col['comp_name']].value,
                                         init_conc=r[col['conc']].value,
                                         final_conc=r[col['final']].value)
                if component.name == diluent:
                    components.append(component)
                    pmix = PCRMasterMix(name=name,
                                        id_=id,
                                        reaction_vol=vol_per_reaction,
                                        sample_vol=sample_vol,
                                        title=title,
                                        components=components,
                                        diluent=component
                                        )
                    line = no_l
                    id = None
                    name = None
                    vol_per_reaction = 25
                    components = []
                    sample_vol = 5  # uL
                    title = None

                elif component.final_conc and (component.id or component.name):
                        components.append(component)

        return pmix


class PCRMasterMixReagent(PreMixReagent):
    """
    Manipulate a PCR Master-Mix Reagent on a robot.
    """

    excess = def_mix_excess

    def __init__(self,
                 pcr_mix: PCRMasterMix,
                 mmix_rack: (lab.Labware, list),
                 num_of_samples: int,
                 pos=None,
                 num_of_aliquots=None,
                 initial_vol=None,
                 def_liq_class=None,
                 excess=None,
                 fill_limit_aliq=None,
                 kit_rack: (lab.Labware, list) = None,
                 primer_mix_rack: (lab.Labware, list) = None,
                 primer_rack: (lab.Labware, list) = None
                 ):
        """
        Construct a robot-usable PCRMasterMixReagent from an abstract PCRMasterMix.
        It is always constructed - no reuse of old aliquots: contains instable components.

        :param primer_mix_rack:
        :param primer_rack:
        :param pos:
        :param num_of_aliquots:
        :param initial_vol:
        :param def_liq_class:
        :param fill_limit_aliq:
        :param pcr_mix:
        :param mmix_rack:
        :param kit_rack:
        :param excess:
        """
        self.pcr_mix = pcr_mix
        components = []
        vol = 0
        lw = mmix_rack[0] if isinstance(mmix_rack, list) else mmix_rack
        reagents = lw.location.worktable.reagents
        diluent_vol = pcr_mix.reaction_vol - pcr_mix.sample_vol
        # num_samples = len(exp.mixes[pcr_mix])
        for component in pcr_mix.components:
            assert isinstance(component, MixComponent)
            if component is pcr_mix.diluent:
                vol_per_reaction = diluent_vol
            else:
                vol_per_reaction = pcr_mix.reaction_vol * component.final_conc / float(component.init_conc)
                diluent_vol -= vol_per_reaction
            if component.name in reagents:
                component_r = reagents[component.name]
                assert isinstance(component_r, Reagent)
                # if not abs(component_r.volpersample - vol_per_reaction) < 0.05:
                #    assert False
                # component_r.put_min_vol()
            else:
                if component.name in PrimerMix.names:

                    for prmx in PrimerMix.names[component.name]:
                        assert isinstance(prmx, PrimerMix)
                        if component.id == prmx.id:
                            primer_mix = prmx
                            break
                    else:
                        raise NoReagentFound(component.name, " as primer mix with ID:" + str(component.id)
                                             + ". Alternatives are: " + str(PrimerMix.names[component.name]))

                    component_r = PrimerMixReagent(primer_mix,
                                                   primer_mix_rack=primer_mix_rack,
                                                   primer_rack=primer_rack)

                else:  # todo for now let consider this a kit component - but it could be a Primer  !!!!
                    component_r = Reagent(component.name,
                                          labware=kit_rack,
                                          volpersample=vol_per_reaction,
                                          num_of_samples=0)
            components.append(component_r)

        PreMixReagent.__init__(self,
                        pcr_mix.name + "-PCRMMix",
                        mmix_rack,
                        components,
                        pos=pos,
                        num_of_aliquots=num_of_aliquots,
                        initial_vol=initial_vol,
                        def_liq_class=def_liq_class,
                        excess=excess or PCRMasterMixReagent.excess,
                        fill_limit_aliq=fill_limit_aliq,
                        num_of_samples=num_of_samples)


class PCReaction:
    """
    Represent abstract information, like an item in some table summarizing reactions in a PCR experiment
    """

    empty = 0
    ntc = 1
    pos = 2
    std = 3
    unk = 4

    rol = {'':empty, None:empty, 'NTC':ntc, 'Pos':pos, 'Unk':unk, 'Std':std}

    def __init__(self,
                 rol,
                 sample=None,
                 targets=None,
                 mix: PCRMasterMix=None,
                 replica=None,
                 row=None,
                 col=None,
                 vol=None):
        self.col = col
        self.row = row
        self.rol = rol
        self.sample = sample
        self.targets = targets or []
        self.mix = mix
        if not mix:
            for target in self.targets:
                if target in PCRMasterMix.names:
                    self.mix = PCRMasterMix.names[target]
                    break
            else:
                assert not self.rol
        self.req_vol = vol or mix
        self.replica = replica

    @staticmethod
    def get_rol(rol):
        rol, replica = rol.split('-')  # todo improve to parse replicas
        return rol in PCReaction.rol, rol, replica

    def __str__(self):
        return self.sample

    def __repr__(self):
        return (self.sample or '-') + '[' + str(self.targets[0] or '-') + ']'


class PCReactionReagent(Reaction):
    vol = 25
    vol_sample = 5

    def __init__(self,
                 pcr_reaction: PCReaction,
                 plate: lab.Labware):
        """

        :param pcr_reaction:
        :param plates:
        """

        pos = plate.Position(pcr_reaction.row, pcr_reaction.col)
        Reaction.__init__(self,
                          name=pcr_reaction.sample + ".PCR" + str(pos),
                          labware=plate,
                          track_sample=pcr_reaction.sample,
                          pos=pos)
        self.pcr_reaction = pcr_reaction
        self.plate = plate


class PCRexperiment:
    """
    Represent abstract information, like an item in some table summarizing PCR experiments
    """

    mixes = {}  #: connect each PCR master mix with the total of well reactions

    def __init__(self,
                 id_ = None,
                 name = None,
                 ncol = 0,
                 nrow = 0):
        """
        A linear rack have just one roe and many columns

        :param id_:
        :param name:
        :param ncol:
        :param nrow:
        """
        self.id = id_
        self.name = name
        self.pcr_reactions = [[PCReaction(PCReaction.empty)]*ncol]*nrow  #: list of PCRReaction to create organized in rows with columns
        self.targets = {}  #: connect each target with a list of PCR reactions well
        self.mixes = {}  #: connect each PCR master mix with a list of well reactions
        self.samples = {}  #: connect each sample with a list of well reactions
        # self.vol = PCReaction.vol
        # self.vol_sample = PCReaction.vol_sample

    def add_reaction(self, pcr_reaction: PCReaction):
        self.pcr_reactions[pcr_reaction.row-1][pcr_reaction.col-1] = pcr_reaction
        self.samples.setdefault(pcr_reaction.sample, []).append(pcr_reaction)
        for target in pcr_reaction.targets:
            self.targets.setdefault(target, []).append(pcr_reaction)
            if target in PCRMasterMix.names:
                mmix = PCRMasterMix.names[target]
                if pcr_reaction.mix:
                    assert pcr_reaction.mix is mmix
                else:
                    pcr_reaction.mix = mmix
        self.mixes.setdefault(pcr_reaction.mix, []).append(pcr_reaction)
        PCRexperiment.mixes.setdefault(pcr_reaction.mix, 0)
        PCRexperiment.mixes[pcr_reaction.mix] += 1
        pass

    def load_excel_list(self, exp_sheet: ExpSheet, load_PCRmix: bool = True):
        if self.id is None:
            self.id = exp_sheet.id
        if self.name is None:
            self.name = exp_sheet.title

        col = {'conc': 16,
               'id': 14,
               'name': 14,
               'vol': 16,
               'final': 18,
               'sample_v': 16,
               'title': 15,
               'comp_name': 15
               }
        logging.debug("opening excel")
        import openpyxl

        if not exp_sheet.file_name:
            exp_sheet.file_name = Path('C:\Prog\exp\PCR fli.xlsx')

        if not exp_sheet.file_name:
            from tkinter import filedialog
            exp_sheet.file_name = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsm"), ("All files", "*.*")),
                                                   defaultextension='fas',
                                                   title='Select HEV isolate subtyping deta')
            if not exp_sheet.file_name:
                return

        if load_PCRmix:
            PCRMasterMix(exp_sheet)

        logging.debug(exp_sheet.file_name)

        wb = openpyxl.load_workbook(str(exp_sheet.file_name))

        logging.debug(wb.sheetnames)

        ws = wb[exp_sheet.page or 'Druken']

        if load_PCRmix:
            PCRMasterMix.load_excel_list(exp_sheet, ws)

        ncol = exp_sheet.num_col
        row = 0
        cell_rows = exp_sheet.cell_rows
        sample_line = exp_sheet.sample_line
        line = 0
        reactions = None

        for r in ws.iter_rows():

            if line == 0:
                line += 1

            elif line == 1:
                reactions = [PCReaction(PCReaction.empty, row=row + 1, col=col + 1) for col in range(ncol)]
                for rx in reactions:
                    rx.rol = PCReaction.rol[r[rx.col].value]
                line += 1

            elif line < sample_line:
                for rx in reactions:
                    target = r[rx.col].value
                    if target:
                        assert rx.rol
                        rx.targets.append(target)
                line += 1
                continue

            elif line == sample_line:
                for rx in reactions:
                    sample = r[rx.col].value
                    if sample:
                        assert rx.rol
                        rx.sample = sample
                line += 1

            elif line <= cell_rows:
                # read cell line
                line += 1

            if line >= cell_rows:  # last cell line
                row += 1
                if len(self.pcr_reactions) < row:
                    self.pcr_reactions += [[None]*ncol]
                for rx in reactions:
                    self.add_reaction(rx)
                line = 1
                if row == exp_sheet.num_row:
                    break

        return self

    def __str__(self):
        return self.name

    def __repr__(self):
        return (self.name or '-') + '[' + str(self.id or '-') + ']'


class PCRexperimentRtic:
    """
    Organize a PCR setup on a robot.
    From a list of abstract information about PCR plate/experiments creates sufficient volume of each of the
    :py:class:`PCRMasterMixReagent` listed in the global PCRexperiment.mixes
    """

    def __init__(self,
                 pcr_exp: (PCRexperiment, list),
                 plates: (lab.Labware, list),
                 kit_rack: (lab.Labware, list),
                 mmix_rack: (lab.Labware, list) = None,
                 primer_mix_rack:  (lab.Labware, list) = None,
                 primer_rack: (lab.Labware, list) = None,
                 protocol=None):
        """


        :param pcr_exp: [:py:class:`PCRexperiment`] abstarct information about the "plate" PCR experiemnts
        :param plates: [:py:class:`labware.Labware`] where to set the PCR reactions.
        :param kit_rack: [racks] in the prefered order to put the PCR kit reagents (stocks solutions)
        :param mmix_rack: [racks] in the prefered order to put the PCR mastermix reagents specially created for these experiments
        :param primer_mix_rack: [racks] in the prefered order to put the primer mix reagents (stocks solutions)s
        :param primer_rack:  [racks] in the prefered order to put the primers reagents (stocks solutions)s
        :param protocol: who invoke this PCR, provide a worktable and the rest of the "environment"
        """
        logging.debug("Creating a PCRexperimentRtic from " + repr(pcr_exp))
        self.pcr_exp = pcr_exp if isinstance(pcr_exp, list) else [pcr_exp]  #: abstract info
        self.plates = plates if isinstance(plates, list) else [plates]
        if len(self.pcr_exp) > len(self.plates):
            raise RuntimeError("Not sufficient plates (", len(self.plates), ") provided for ",
                               len(self.pcr_exp), " PCR experiments.")
        self.protocol = protocol
        self.mixes = {}  #: connect each :py:class:`PCRMasterMix` in the experiment with the PCR wells into which will be pippeted
        mixes = {}  # connect each PCRMasterMix with the corresponding PCRMasterMixReagent
        for pcr_mix, pcr_reactions in PCRexperiment.mixes.items():
            if not pcr_mix:  # empty reaction wells are market with None mix.
                continue
            assert isinstance(pcr_mix, PCRMasterMix)
            mix = PCRMasterMixReagent(pcr_mix=pcr_mix,
                                      mmix_rack=mmix_rack,
                                      num_of_samples=pcr_reactions,
                                      primer_mix_rack=primer_mix_rack,
                                      kit_rack=kit_rack,
                                      primer_rack=primer_rack)
            self.mixes[mix] = []
            mixes[pcr_mix] = mix

        for exp, plate in zip(self.pcr_exp, self.plates):  # iterate each PCR plate for which we have PCR mixes declared
            assert isinstance(exp, PCRexperiment)
            for pcr_mix, pcr_reactions in exp.mixes.items():  # visit each PCR mix with corresponding list of reactions
                if not pcr_mix:  # empty reaction wells are market with None mix.
                    continue
                assert isinstance(pcr_mix, PCRMasterMix)

                mix = mixes[pcr_mix]
                sv = pcr_mix.sample_vol  # all reactions prepared with this mix have the same reaction and sample volume
                nw = len(pcr_reactions)
                react_wells = []
                for rx in pcr_reactions:
                    r = PCReactionReagent(rx, plate)
                    react_wells.extend(r.Replicas)
                self.mixes.setdefault(mix, []).extend(react_wells)                     # just samples?
                pass

        pass

    def pippete_mix(self):
        pass

    def pippete_samples(self):
        pass

    def dilute_primers(self):
        for reagent in self.protocol.reagents:
            if isinstance(reagent, PrimerReagent):
                if not reagent.primer.prepared:
                    reagent.dilute()

    def vol (self, vol, vol_sample):
        self.vol = vol
        self.vol_sample = vol_sample

    def __str__(self):
        return "PCR Exp " + str(self.pcr_exp)

    def __repr__(self):
        return str(self)  # todo review


if __name__ == '__main__':
    logging.getLogger(__name__).setLevel(10)

    primers = Primer.load_excel_list()
    primermixes = PrimerMix.load_excel_list()

    sheet0 = ExpSheet(file_name=Path('K:\AG RealtimePCR\Ariel\Exp 424. WESSV.MID.NewRNAbis-4. AVRvsSAfr.PanFlav-224.Ute.xlsx'),
                      page='Druken (2)',
                      cell_rows=3,
                      sample_line=3)

    sheet1 = ExpSheet(file_name=Path('C:\\Users\\Ariel\\Documents\\Exp\\PCR\\Exp 308. WNV.ZKU.10-1 10-10. WN-INNT-133, WN.Hoff, PanFlav.116.pltd.xlsx'),
                      page='Druken (3)',
                      cell_rows=6,
                      sample_line=6)
    # pcrmixes = PCRMasterMix.load_excel_list(sheet)
    exp = PCRexperiment().load_excel_list(sheet0)

    pass

